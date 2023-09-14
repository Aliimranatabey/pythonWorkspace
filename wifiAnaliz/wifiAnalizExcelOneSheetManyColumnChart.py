from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from datetime import datetime

def move_dashboard_to_front(file_path, dashboard_name):
    wb = load_workbook(file_path)
    dashboard = wb[dashboard_name]
    # "Dashboard" sayfasının indeksini al
    dashboard_index = wb.sheetnames.index(dashboard.title)
    # "Dashboard" sayfasını en başa taşı
    wb._sheets.insert(0, wb._sheets.pop(dashboard_index))
    # Dosyayı kaydet
    wb.save(file_path)

def create_combined_chart(master_file):
    wb = load_workbook(master_file)
    data_ws = wb['All Sheets']  # Verilerin olduğu sheet
    chart_ws = wb.create_sheet(title='Dashboard')  # Yeni bir sheet oluştur

    def add_chart_to_dashboard(chart, row, col):
        chart_ws.add_chart(chart, f"{chr(65+col)}{row+1}")
    time_values = []    
    for row in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, min_col=1, max_col=1):
        cell_j = row[0]
        cell_value = cell_j.value
        datetime_obj = datetime.strptime(cell_value.strip(), '%a %b %d %H:%M:%S %Y')
        time_str = datetime_obj.strftime('%H:%M:%S')
        time_values.append(time_str)

    time_values_ref = Reference(data_ws, min_col=10, min_row=2, max_row=data_ws.max_row, max_col=10)
    for i, time_value in enumerate(time_values):
        data_ws.cell(row=i+2, column=10).value = time_value

    # State Chart
    state_chart = LineChart()
    state_chart.title = "State Chart"
    state_chart.y_axis.title = "State Value"
    state_chart.x_axis.title = "Time"
    state_data = Reference(data_ws, min_col=3, min_row=2, max_col=3, max_row=data_ws.max_row)
    # state_cats = Reference(data_ws, min_col=10, min_row=2, max_row=data_ws.max_row, max_col=10)
    state_chart.add_data(state_data, titles_from_data=True)
    state_chart.set_categories(time_values_ref)
    state_chart.y_axis.majorGridlines = None
    state_chart.y_axis.majorUnit = 1
    state_chart.y_axis.scaling.max = 2
    state_chart.width = state_chart.width * 2.5
    state_chart.height = state_chart.height * 2
    for row in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, min_col=3, max_col=9):
        cell_b = row[2]
        if state_chart.title == "State Chart":
            if cell_b.value == '1':
                cell_b.value = 1
            elif cell_b.value == '0':
                cell_b.value = 0
            else:
                cell_b.value = 2
    add_chart_to_dashboard(state_chart, 0, 0)
#     # Son olarak, yeni oluşturulan sheet'i kaydet
    
    signal_chart = LineChart()
    signal_chart.title = "Signal Chart"
    signal_chart.y_axis.title = "Signal Value (%)"
    signal_chart.x_axis.title = "Time"
    signal_data = Reference(data_ws, min_col=4, min_row=2, max_col=4, max_row=data_ws.max_row)
    # signal_cats = Reference(data_ws, min_col=1, min_row=2, max_row=data_ws.max_row)
    signal_chart.add_data(signal_data, titles_from_data=True)
    signal_chart.set_categories(time_values_ref)
    signal_chart.y_axis.majorGridlines = None
    signal_chart.y_axis.scaling.min = 0
    signal_chart.y_axis.scaling.max = 100
    signal_chart.y_axis.majorUnit = 10
    signal_chart.width = signal_chart.width * 2.5
    signal_chart.height = signal_chart.height * 2

    for row in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, min_col=4, max_col=9):
        cell_d = row[3]  # D sütunu
        for i in range(1,101):
            if cell_d.value == f'{i}':
                cell_d.value = i
    add_chart_to_dashboard(signal_chart, 32, 0)

    tr_rr_chart = LineChart()
    tr_rr_chart.title = "Tr Rr Chart"
    tr_rr_chart.y_axis.title = "Value"
    tr_rr_chart.x_axis.title = "Time"
    rr_data = Reference(data_ws, min_col=5, min_row=2, max_col=5, max_row=data_ws.max_row)
    tr_data = Reference(data_ws, min_col=6, min_row=2, max_col=6, max_row=data_ws.max_row)
    # tr_rr_cats = Reference(data_ws, min_col=1, min_row=2, max_row=data_ws.max_row)
    tr_rr_chart.add_data(tr_data, titles_from_data=True)
    tr_rr_chart.add_data(rr_data, titles_from_data=True)
    tr_rr_chart.set_categories(time_values_ref)
    tr_rr_chart.y_axis.majorGridlines = None
    tr_rr_chart.y_axis.scaling.min = 0
    tr_rr_chart.y_axis.scaling.max = 1000
    tr_rr_chart.y_axis.majorUnit = 100
    tr_rr_chart.width = tr_rr_chart.width * 2.5
    tr_rr_chart.height = tr_rr_chart.height * 2
# Combined Chart oluştur
    add_chart_to_dashboard(tr_rr_chart, 64, 0)

    channel_chart = LineChart()
    channel_chart.title = "Channel Chart"
    channel_chart.y_axis.title = "Value"
    channel_chart.x_axis.title = "Time"
    channel_data = Reference(data_ws, min_col=7, min_row=2, max_col=7, max_row=data_ws.max_row)
    # channel_cats = Reference(data_ws, min_col=1, min_row=2, max_row=data_ws.max_row)
    channel_chart.add_data(channel_data, titles_from_data=True)
    channel_chart.set_categories(time_values_ref)
    channel_chart.y_axis.majorGridlines = None
    channel_chart.y_axis.majorUnit = 1
    channel_chart.y_axis.scaling.max = 24
    channel_chart.y_axis.scaling.min = 0
    channel_chart.width = channel_chart.width * 2.5
    channel_chart.height = channel_chart.height * 2
    for row in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, min_col=1, max_col=9):
        cell_f = row[6]  # G sütunu
        for i in range(1,25):
            if cell_f.value == f'{i}':
                cell_f.value = i
    add_chart_to_dashboard(channel_chart, 96, 0)

    protocol_chart = LineChart()
    protocol_chart.title = "Protocol Chart"
    protocol_chart.y_axis.title = "Value"
    protocol_chart.x_axis.title = "Time"
    protocol_data = Reference(data_ws, min_col=9, min_row=2, max_col=9, max_row=data_ws.max_row)
    # protocol_cats = Reference(data_ws, min_col=1, min_row=2, max_row=data_ws.max_row)
    protocol_chart.add_data(protocol_data, titles_from_data=True)
    protocol_chart.set_categories(time_values_ref)
    protocol_chart.y_axis.majorGridlines = None
    protocol_chart.y_axis.majorUnit = 1
    protocol_chart.y_axis.scaling.max = 8
    protocol_chart.width = protocol_chart.width * 2.5
    protocol_chart.height = protocol_chart.height * 2
    for row in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, min_col=1, max_col=9):
        cell_i = row[8]  # I sütunu
        if protocol_chart.title=="Protocol Chart":
            if cell_i.value == '5':
                cell_i.value = 5
            elif cell_i.value == '4':
                cell_i.value = 4
            elif cell_i.value == '3':
                cell_i.value = 3
            elif cell_i.value == '2':
                cell_i.value = 2
            else:
                cell_i.value = 1
    add_chart_to_dashboard(protocol_chart, 127, 0)
    for row in data_ws.iter_rows(min_row=2, max_row=data_ws.max_row, min_col=1, max_col=10):
        cell_j=row[9]
        cell_j.value=""
    wb.save(master_file)
create_combined_chart("onesheet_master_sheet.xlsx")
move_dashboard_to_front('onesheet_master_sheet.xlsx', 'Dashboard')
